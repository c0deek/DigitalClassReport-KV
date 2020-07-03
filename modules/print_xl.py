import openpyxl
import datetime
import json
import os

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

from django.conf import settings

def generate_xl(data, teachers):

	# parse data and teachers
	data = json.loads(data)
	teachers = json.loads(teachers)
	
	wb = Workbook()

	sheet = wb.active

	sheet.title = "Custom title"

	fontH = Font(name = 'Montserrat',
				 size = 15,
				 bold = True,
				)
				
	fontT = Font(name = 'Montserrat',
				 size = 8,
				 bold = True,
				)
				
	fontD = Font(name = 'Montserrat',
				 size = 8,
				)
				
	align = Alignment(horizontal = 'center',
					  vertical = 'center',
					 )
					 
	border = Border(right=Side(border_style='thin',
								  color='121212'),
								  
					bottom=Side(border_style='thin',
								    color='121212'),
					)

	columns = "A B C D E F G H I J K L".split(" ")

	sheet.merge_cells('A1:L2')

	sheet['A1'] = f'Kendriya Vidyalaya No.2, Armapur, Kanpur\n E-Learning Class Completion Report Date: {data[0]["fields"]["date"][5:] + "-" + data[0]["fields"]["date"][:4]}'
	
	sheet['A3'] = "Name of the Teacher"
	sheet['B3'] = "Class"
	sheet['C3'] = "Subject"
	sheet['D3'] = "Total Students"
	sheet['E3'] = "Present"
	sheet['F3'] = "Absent"
	sheet['G3'] = "Topic"
	sheet['H3'] = "Platform Used"
	sheet['I3'] = "Homework"
	sheet['J3'] = "Observation"
	sheet['K3'] = "Remarks"
	sheet['L3'] = "Observed By"

	i = 3
	for entry in data:
		i += 1
		teacher_pk = entry["fields"]['teacher']
		for teacher_entry in teachers:
			if(teacher_entry["pk"] == teacher_pk):
				teacher_pk = teacher_entry["fields"]["name"]
				break;
				
				
		sheet['A'+str(i)] = teacher_pk
		sheet['B'+str(i)] = str(entry["fields"]['Class']) + " " + entry["fields"]['section']
		sheet['C'+str(i)] = entry["fields"]['subject']
		sheet['D'+str(i)] = entry["fields"]['total_students']
		sheet['E'+str(i)] = entry["fields"]['present']
		sheet['F'+str(i)] = entry["fields"]['total_students'] - entry["fields"]['present']
		sheet['G'+str(i)] = entry["fields"]['topic']
		sheet['H'+str(i)] = entry["fields"]['platform']
		sheet['I'+str(i)] = entry["fields"]['homework']
		sheet['J'+str(i)] = entry["fields"]['observation']
		sheet['K'+str(i)] = entry["fields"]['remark']
		sheet['L'+str(i)] = entry["fields"]['observed_by']


        for col in range(1, len(columns)+1):
            sheet.column_dimensions['A'].width = 15
            sheet.column_dimensions['B'].width = 5
            sheet.column_dimensions['C'].width = 10
            sheet.column_dimensions['D'].width = 9
            sheet.column_dimensions['E'].width = 6
            sheet.column_dimensions['F'].width = 6
            sheet.column_dimensions['G'].width = 12
            sheet.column_dimensions['H'].width = 16
            sheet.column_dimensions['I'].width = 13
            sheet.column_dimensions['L'].width = 14
            sheet.column_dimensions['J'].width = 9
            sheet.column_dimensions['K'].width = 12
            sheet.column_dimensions['L'].width = 8
            
            for row in range(1, 4 + len(data)):
                cell = sheet.cell(row = row, column = col)
                if(row == 1):
                    cell.font = fontH
                    sheet.row_dimensions[row].height = 25
                
                elif(row == 3):
                    cell.font = fontT
			
                else:
                    cell.font = fontD
			
                    cell.alignment = align
                    cell.border = border
	
        file_path = os.path.join(settings.MEDIA_ROOT, 'report.xlsx')
        print("Excel Creater", file_path)
        wb.save(filename = file_path)
